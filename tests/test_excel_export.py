"""Tabular data export (v0.61): the Total Envelope QA package - Settings
sheet, one unfactored sheet per validly applied load case, one Total
Envelope sheet per analyzed result mode - and its CSV twin."""
import io

import numpy as np
import openpyxl
import pandas as pd
import pytest

import bricos_data as data
import bricos_export as export
import bricos_solver as solver


def _run(params):
    solver.clear_solver_cache()
    raw, nodes, _, err = solver.run_raw_analysis(params)
    assert err == 0
    solver.clear_solver_cache()
    return raw


def _beam_params(name="Sys", **overrides):
    params = data.get_clear("A", "Beam")
    params.update({
        "name": name,
        "mode": "Beam",
        "num_spans": 2,
        "L_list": [10.0] * 10,
        "Is_list": [0.8] * 10,
        "sw_list": [5.0, 5.0] + [0.0] * 8,
        "E": 30e6,
        "E_span_list": [30e6] * 10,
        "mesh_size": 0.5,
        "step_size": 0.5,
        "phi_mode": "Manual",
        "phi": 1.0,
        "vehicle": {"loads": [10.0], "spacing": [0.0]},
        "vehicle_direction": "Forward",
        "udl_q": 2.5,
        "udl_gap": 2.0,
        "udl_mode": "Moving",
        "KFI": 1.1,
        "gamma_g": 1.0,
        "gamma_veh": 1.4,
        "gamma_udl": 0.56,
        "analyze_uls": True,
        "analyze_sls": True,
    })
    params.update(overrides)
    return params


# --- presence semantics ---

def test_system_has_case_checks_each_load_case():
    p = _beam_params()
    assert export.system_has_case(p, "Dead Load")
    assert export.system_has_case(p, "Traffic UDL")
    assert export.system_has_case(p, "Vehicle Envelope")
    assert not export.system_has_case(p, "Soil")
    assert not export.system_has_case(p, "Surcharge")

    p_none = _beam_params(sw_list=[0.0] * 10, udl_q=0.0,
                          vehicle={"loads": [], "spacing": []})
    assert not export.system_has_case(p_none, "Dead Load")
    assert not export.system_has_case(p_none, "Traffic UDL")
    assert not export.system_has_case(p_none, "Vehicle Envelope")


# --- sheet assembly ---

def test_total_envelope_export_sheets_match_applied_cases_and_modes():
    pA = _beam_params("Alpha")
    raw = _run(pA)
    settings, sheets = export.total_envelope_export(pA, raw, _beam_params(), None, False)

    assert list(sheets.keys()) == [
        "Dead Load (Unfactored)",
        "Traffic UDL (Unfactored)",
        "Vehicle Envelope (Unfactored)",
        "Total Envelope (ULS)",
        "Total Envelope (SLS)",
        "Total Envelope (Unfactored)",
    ]
    # No Soil/Surcharge sheets: not defined for any system.
    assert "Parameter" in settings.columns
    assert "System A" in settings.columns
    assert "System B" not in settings.columns


def test_total_envelope_export_respects_limit_state_toggles():
    pA = _beam_params(analyze_uls=True, analyze_sls=False)
    raw = _run(pA)
    _, sheets = export.total_envelope_export(pA, raw, _beam_params(), None, False)

    assert "Total Envelope (ULS)" in sheets
    assert "Total Envelope (SLS)" not in sheets
    assert "Total Envelope (Unfactored)" in sheets  # always available


def test_one_sided_load_case_exports_only_the_loaded_system():
    pA = _beam_params("Alpha")
    pB = _beam_params("Beta", udl_q=0.0)
    raw_A = _run(pA)
    raw_B = _run(pB)

    _, sheets = export.total_envelope_export(pA, raw_A, pB, raw_B, True)

    udl = sheets["Traffic UDL (Unfactored)"]
    assert set(udl["System"]) == {"Alpha"}
    # Both systems appear where both carry the case.
    assert set(sheets["Dead Load (Unfactored)"]["System"]) == {"Alpha", "Beta"}


def test_envelope_sheet_values_equal_combined_results():
    # v0.62: the repeated section x at every mesh sub-element boundary is
    # merged to one row (max of maxima / min of minima over the repeats),
    # so compare each exported row against the raw arrays at its x.
    pA = _beam_params("Alpha")
    raw = _run(pA)
    _, sheets = export.total_envelope_export(pA, raw, _beam_params(), None, False)

    res_uls = solver.combine_results(raw, pA, "Design (ULS)")
    te = res_uls["Total Envelope"]["S1"]
    x_raw = np.asarray(te["x"])
    s1 = sheets["Total Envelope (ULS)"].query("Element == 'S1'")

    assert s1["Location [m]"].is_unique
    np.testing.assert_allclose(
        s1["Location [m]"].to_numpy(), np.unique(x_raw), atol=1e-12)
    for _, row in s1.iterrows():
        at_x = np.isclose(x_raw, row["Location [m]"], rtol=0.0, atol=1e-9)
        assert row["M_max [kNm]"] == pytest.approx(
            np.max(np.asarray(te["M_max"])[at_x]), abs=1e-9)
        assert row["M_min [kNm]"] == pytest.approx(
            np.min(np.asarray(te["M_min"])[at_x]), abs=1e-9)
        # Deflections are exported in mm.
        assert row["Def_Y_min [mm]"] == pytest.approx(
            np.min(np.asarray(te["def_y_min"])[at_x]) * 1000, abs=1e-9)


def test_duplicate_boundary_sections_merge_conservatively():
    # Two sub-elements sharing x=1.0 with slightly different bounds (the
    # segment-snapped UDL window case): one row survives, carrying the
    # max of the maxima and the min of the minima.
    r_dict = {"S1": {
        "x": np.array([0.0, 1.0, 1.0, 2.0]),
        "M_max": np.array([1.0, 10.0, 10.5, 3.0]),
        "M_min": np.array([-1.0, -8.0, -8.4, -2.0]),
        "V_max": np.zeros(4), "V_min": np.zeros(4),
        "N_max": np.zeros(4), "N_min": np.zeros(4),
        "def_x_max": np.zeros(4), "def_x_min": np.zeros(4),
        "def_y_max": np.zeros(4), "def_y_min": np.zeros(4),
    }}
    df = export.case_dataframe(r_dict, "Sys")

    assert list(df["Location [m]"]) == [0.0, 1.0, 2.0]
    merged = df.set_index("Location [m]").loc[1.0]
    assert merged["M_max [kNm]"] == pytest.approx(10.5)
    assert merged["M_min [kNm]"] == pytest.approx(-8.4)


def test_step_mode_keeps_physical_jumps_and_drops_pure_repeats():
    base = {
        "x": np.array([0.0, 1.0, 1.0, 2.0, 2.0, 3.0]),
        # Pure repeat at x=1.0; a real shear jump (axle) at x=2.0.
        "M": np.array([0.0, 5.0, 5.0, 4.0, 4.0, 0.0]),
        "V": np.array([2.0, 2.0, 2.0, 2.0, -8.0, -8.0]),
        "N": np.zeros(6),
        "def_x": np.zeros(6), "def_y": np.zeros(6),
    }
    df = export.case_dataframe({"S1": base}, "Sys", step_mode=True)

    assert list(df["Location [m]"]) == [0.0, 1.0, 2.0, 2.0, 3.0]
    jump = df[df["Location [m]"] == 2.0]["V [kN]"].tolist()
    assert jump == [pytest.approx(2.0), pytest.approx(-8.0)]


def test_exported_sheets_have_unique_locations_per_element():
    pA = _beam_params("Alpha")
    raw = _run(pA)
    _, sheets = export.total_envelope_export(pA, raw, _beam_params(), None, False)

    for name, df in sheets.items():
        dup = df.duplicated(subset=["System", "Element", "Location [m]"])
        assert not dup.any(), f"duplicated section rows in {name}"


# --- settings sheet ---

def test_settings_never_leak_raw_no_kfi_soil_value():
    pA = _beam_params(KFI=1.1, gamma_j=1.0 / 1.1)
    raw = _run(pA)
    settings, _ = export.total_envelope_export(pA, raw, _beam_params(), None, False)

    row = settings.loc[settings["Parameter"] == "γ_j - Soil (ULS)", "System A"]
    assert row.item() == "1.0 (No KFI)"
    blob = settings.to_csv(index=False)
    assert "0.909" not in blob


def test_settings_state_factors_and_vehicle_definition():
    pA = _beam_params()
    raw = _run(pA)
    settings = export.settings_dataframe(pA, _beam_params(), raw, None, False)

    by_param = dict(zip(settings["Parameter"], settings["System A"]))
    assert by_param["KFI (consequence class)"] == 1.1
    assert by_param["γ_veh,A - Vehicle A (ULS)"] == 1.4
    assert by_param["γ_UDL - Traffic UDL (ULS)"] == 0.56
    assert by_param["Traffic UDL q [kN/m]"] == pytest.approx(2.5)
    assert "loads [t]: 10" in by_param["Vehicle A"]
    assert by_param["Vehicle B"] == "none"
    assert "clear distance 2 m" in by_param["Traffic UDL application"]
    assert by_param["Limit states analyzed"] == "ULS and SLS"


# --- serialization ---

def test_xlsx_roundtrip_preserves_sheets_and_values():
    pA = _beam_params("Alpha")
    raw = _run(pA)
    settings, sheets = export.total_envelope_export(pA, raw, _beam_params(), None, False)

    blob = export.to_xlsx_bytes(settings, sheets)
    wb = openpyxl.load_workbook(io.BytesIO(blob))

    assert wb.sheetnames == ["Settings"] + list(sheets.keys())
    ws = wb["Total Envelope (ULS)"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header[:3] == ["System", "Element", "Location [m]"]
    # First data row matches the source frame.
    first = sheets["Total Envelope (ULS)"].iloc[0]
    row2 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert row2[0] == "Alpha" and row2[1] == first["Element"]
    assert row2[3] == pytest.approx(first["M_max [kNm]"])

    ws_set = wb["Settings"]
    params_col = [r[0].value for r in ws_set.iter_rows(min_col=1, max_col=1)]
    assert "BriCoS version" in params_col and "KFI (consequence class)" in params_col


def test_csv_package_carries_settings_block_and_tagged_results():
    pA = _beam_params("Alpha")
    raw = _run(pA)
    settings, sheets = export.total_envelope_export(pA, raw, _beam_params(), None, False)

    text = export.to_csv_bytes(settings, sheets).decode("utf-8-sig")
    settings_block, results_block = text.split("\n\n", 1)

    head = pd.read_csv(io.StringIO(settings_block))
    assert list(head.columns) == ["Parameter", "System A"]

    body = pd.read_csv(io.StringIO(results_block))
    assert body.columns[0] == "Result set"
    assert set(body["Result set"]) == set(sheets.keys())
    uls_rows = body[body["Result set"] == "Total Envelope (ULS)"]
    assert len(uls_rows) == len(sheets["Total Envelope (ULS)"])


# --- detailed frame builder (shared with the Tabular Data view) ---

def test_case_dataframe_step_mode_and_envelope_mode_columns():
    pA = _beam_params()
    raw = _run(pA)
    res = solver.combine_results(raw, pA, "Unfactored")

    env = export.case_dataframe(res["Total Envelope"], "Sys")
    assert "M_max [kNm]" in env.columns and "Def_Y_min [mm]" in env.columns

    step = export.case_dataframe(res["Vehicle Steps A"][0]["res"], "Sys", step_mode=True)
    assert "M [kNm]" in step.columns and "M_max [kNm]" not in step.columns

    assert export.case_dataframe({}, "Sys") is None
